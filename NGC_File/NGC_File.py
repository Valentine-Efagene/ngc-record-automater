from openpyxl import load_workbook, Workbook
import pickle
import os
import time
import threading
from datetime import datetime

cellNumberPath = 'cellNumber'

def saveCellNumber(cellNumber):
    try:
        p_out = open(cellNumberPath, 'wb')
        pickle.dump(cellNumber, p_out)
        p_out.close()
    except:
        print 'Could not save cell number'

def readCellNumber():
    try:
        p_in = open(cellNumberPath, 'rb')
        result = pickle.load(p_in)
        return result
    except EOFError as error:
        print 'Could not write cell number'

def monitorFolder(path):
    list_of_files = os.listdir(path)
    length = len(list_of_files)
    wb = load_workbook("minutes.xlsx")
    cellNumber = readCellNumber()

    while(True):
        list_of_files = os.listdir(path)
        paths = [os.path.join(path, basename) for basename in list_of_files]

        if (len(list_of_files) > 0) and len(list_of_files) > length:
            latest_file = max(paths, key=os.path.getctime)
            print os.path.basename(latest_file)
            length = len(list_of_files)
            # grab the active worksheet
            ws = wb.active

            # Data can be assigned directly to cells
            ws[nameLoc+str(cellNumber)] = os.path.basename(latest_file)
            ws[dateLoc+str(cellNumber)] = datetime.fromtimestamp(os.path.getctime(latest_file)).strftime('%d-%m-%y')
            ws[deptLoc+str(cellNumber)] = os.path.basename(os.path.abspath(os.path.join(latest_file, os.pardir)))
            ws[timeLoc+str(cellNumber)] = datetime.fromtimestamp(os.path.getctime(latest_file)).strftime('%H:%M:%S')
            ws[pathLoc+str(cellNumber)].hyperlink = latest_file
            ws.value = latest_file
            ws.style = "Hyperlink"
            cellNumber += 1
            saveCellNumber(cellNumber)
            # Save the file
            wb.save("minutes.xlsx")

if not os.path.exists(cellNumberPath):
    saveCellNumber(2)

nameLoc = 'A'
dateLoc = 'B'
deptLoc = 'C'
timeLoc = 'D'
pathLoc = 'E'
minuteLoc = 'F'
cellNumber = readCellNumber()

if not os.path.exists('minutes.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws[nameLoc+str(1)] = 'FILE NAME'
    ws[dateLoc+str(1)] = 'DATE'
    ws[dateLoc+str(1)] = 'DEPARTMENT'
    ws[timeLoc+str(1)] = 'TIME'
    ws[pathLoc+str(1)] = 'LINK'
    ws[minuteLoc+str(1)] = 'MINUTE'
    wb.save("minutes.xlsx")

with open('folders') as fp:
        WatchDirectories = fp.read().split("\n")



for f in WatchDirectories:
    if f == '':
        continue

    t = threading.Thread(target = monitorFolder, args=(f,))
    t.start()